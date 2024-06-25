import pandas as pd
import openpyxl
import time
import dbConnectSQL
import os
from dotenv import load_dotenv, dotenv_values 

start = time.time()
dataframe = openpyxl.load_workbook("testxls.xlsm")
db = dbConnectSQL.DbConnectSQL("config1.txt")
load_dotenv()

dataframe1 = dataframe.active
mydata = []
dataset = []

for row in range(0, dataframe1.max_row):
    mydata.append([col[row].value for col in dataframe1.iter_cols(1, dataframe1.max_column)])

for i in range(len(mydata[0])):
    dataset.append([mydata[j][i] for j in range(len(mydata))])


for i in range(len(dataset)-1):
    s = set(dataset[i])
    for el in s:
        str1 = str(el).replace("megafon\\","").rstrip() + "@megafon.ru"
        val = "%" + str(str1) + "%"
        sql = "SELECT password FROM megafon_22.user WHERE login LIKE '" + val + "'"
        
        db.mycursor.execute(sql)
        rowdata = db.mycursor.fetchall()
        # print(str1)
        if len(rowdata) != 0:
            if rowdata[0] == os.getenv('pass'): print(str1, ' ', rowdata[0])
        if len(rowdata) == 0: print(str1)
        checkSet = list(filter(lambda x: str(x).find(str1) != -1, dataset[4]))
        if len(checkSet) > 0: print(el)
    # for j in range(len(dataset[i])):
    #     str1 = str(dataset[i][j]).replace("megafon\\","").rstrip() + "@megafon.ru"
    #     checkSet = list(filter(lambda x: str(x).find(str1) != -1, dataset[4]))
    #     if len(checkSet) > 0: print(str1, j)
        
end = time.time()
print("The time : ",round(end - start,5))